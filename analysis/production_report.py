#%%

#region importacoes
import psycopg2
from dotenv import load_dotenv
import os
from psycopg2.extras import DictCursor
import pandas as pd
import numpy as np
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from pathlib import Path
from datetime import datetime
from babel.numbers import format_currency

#endregion

#%%
#region funcoes
def get_connect_using_psycopg2(
      dbname, user, password, host, port, sslmode = 'disable'
):



    # conn = psycopg2.connect(
    #     dbname=os.getenv('DB_NAME'),
    #     user=os.getenv('DB_USER'),
    #     password=os.getenv('DB_PASSWORD'),
    #     host=os.getenv('DB_HOST'),
    #     port=os.getenv('DB_PORT'),
    #     sslmode='disable'
    # )

    conn = psycopg2.connect(
        dbname=dbname,
        user=user,
        password=password,
        host=host,
        port=port,
        sslmode=sslmode
    )

    return conn

def execute_query(dbname, user, password, host, port, query):

    conn = get_connect_using_psycopg2(
       dbname,user,password,host,port
    )

    cursor = conn.cursor(cursor_factory=DictCursor)

    cursor.execute(query)

    data = cursor.fetchall()

    conn.close()

    results_as_dict = [dict(result) for result in data]

    return results_as_dict

def converter_para_float(valor):
    try:
        return float(valor)
    except:
        return np.nan

def carregar_atividades_colaboradores(
      dbname, user, password, host, port, initialDate, finalDate, contractId, colaboradorId = None
):
  sql = """
    select u.id as "ColaboradorId", u.nome as "Colaborador" , c.titulo as "Cargo", r."data" as "Data" , a.descricao as "Atividade", f.nome as "Furo", ar.quantidade as "Quantidade"
    from atividade_rdos ar
    inner join rdo_users ru on ru.rdo_id  = ar.rdo_id
    inner join rdos r on r.id = ar.rdo_id
    inner join atividades a on a.id = ar.atividade_id
    inner join furos f on f.id = ar.furo_id
    inner join users u on u.id = ru.user_id
    inner join cargos c on c.id  = u.cargo_id
    where a.tipo  = 'produtiva'
    {condicao_colaborador}
    and r."data" between '{initialDate}' and '{finalDate}'
    and r.contrato_id = {contractId}
    and f.invalid  is false
    order by u.nome, r.data, a.descricao , f.nome
    """.format(
      initialDate=initialDate,
      finalDate=finalDate,
      contractId=contractId,
      condicao_colaborador = f"and u.id = {colaboradorId}" if colaboradorId is not None else ""
    )



  data = execute_query(dbname, user, password, host, port, sql)

  df = pd.DataFrame(data)

  return df


def carregar_dados_resumo_memoria(
      dbname,
      user,
      password,
      host,
      port,
      initialDate,
      finalDate,
      contractId,
):



    def carregar_do_banco():
      sql = """
          select u.id as "ColaboradorId", u.nome as "Colaborador" , r.data as "Data" , c.titulo as "Cargo" , a.descricao as "Atividade" , f.nome as "Furo" , ar.quantidade as "Quantidade" , acv.valor_unitario as "Valor Unitário"
          from atividade_rdos ar
          inner join rdo_users ru on ru.rdo_id  = ar.rdo_id
          inner join rdos r on r.id = ar.rdo_id
          inner join atividades a on a.id = ar.atividade_id
          inner join furos f on f.id = ar.furo_id
          inner join users u on ru.user_id  = u.id
          inner join cargos c on c.id = u.cargo_id
          inner join atividade_cargo_valores acv on acv.cargo_id  = u.cargo_id and acv.atividade_id = ar.atividade_id
          where r."data" between '{inicio}' and '{fim}'
          and a.tipo  = 'produtiva'
          and r.contrato_id = {contrato}
          and f.invalid is false
          order by u.nome , r."data" , a.descricao , f.nome
      """.format(
          inicio=initialDate,
          fim=finalDate,
          contrato=contractId
      )

      data = execute_query(dbname, user,password,host,port, sql)

      df = pd.DataFrame(data)



      return df

    df = carregar_do_banco()

    if df.empty:
      return df


    df['Colaborador'] = df['Colaborador'].apply(lambda x: x.strip())
    df['Quantidade'] = df['Quantidade'].astype(float)
    df['Valor Unitário'] = df['Valor Unitário'].astype(float)

    colunas_indesejadas = [
      'Mudança de Sonda (Maior que 200m)',
      'Mudança de Sonda (Menor que 200m)',
      'Mudança de Sonda (Menor que 200 m)',
      'Mudança de Sonda (Menor que 200 m)',
      'Mudança de sonda (Km adicional)',
      'Tamponamento Furo Mts',
      'Tamponamento de furo'
    ]

    df = df[~df['Atividade'].isin(colunas_indesejadas)]


    return df

def obter_limite_pagamento(colaborador_id, contrato_id, dbname, user, password, host, port):
  def obter_cargo_id():
     sql = f"""
       select cargo_id from users where id = {colaborador_id}
      """

     data = execute_query(dbname,user,password,host,port,sql)

     return data[0]['cargo_id']

  def obter_limite_pagamento_por_cargo(cargo_id):
    sql = f"""
    select limite from limite_pagamentos where contrato_id = {contrato_id} and cargo_id = {cargo_id}
    """

    data = execute_query(dbname,user,password,host,port,sql)

    return float(data[0]['limite']) if len(data) > 0 else None

  limite_pagamento = obter_limite_pagamento_por_cargo(
     cargo_id=obter_cargo_id()
  )

  return limite_pagamento

def gerar_memoria_de_calculo_por_colaborador(df, nome_colaborador, dbname, user, password, host, port, contrato_id):

  df_colaborador = df[df['Colaborador'] == nome_colaborador]

  df_colaborador_agregado = df_colaborador.pivot_table(
      index = ['ColaboradorId', 'Colaborador', 'Cargo', 'Atividade'],
      values=['Quantidade', 'Valor Unitário'],
      aggfunc={
          'Quantidade': 'sum',
          'Valor Unitário': 'first'
      }
  )

  df_colaborador_agregado['Subtotal'] = df_colaborador_agregado['Quantidade'] * df_colaborador_agregado['Valor Unitário']


  def gerar_total_geral(df):
      s = pd.Series(np.nan, index=df.index)

      s.iloc[0] = df['Subtotal'].sum()

      return s




  df_colaborador_agregado['Valor a pagar'] = gerar_total_geral(df_colaborador_agregado)

  df_colaborador_agregado['Valor a pagar'] = df_colaborador_agregado['Valor a pagar'].apply(converter_para_float).apply(lambda x : round(x,2))

  limite_series = pd.Series(np.nan, index=df_colaborador_agregado.index)

  limite_series.iloc[0] = obter_limite_pagamento(colaborador_id=df_colaborador_agregado.reset_index()['ColaboradorId'].iloc[0],
     contrato_id=contrato_id,
     dbname=dbname, user=user, password=password, host=host, port=port)

  df_colaborador_agregado['Limite'] = limite_series

  # df_colaborador_agregado = df_colaborador_agregado.reset_index().drop(['CargoId'],axis=1).set_index(['ColaboradorId','Colaborador','Cargo','Atividade'])

  return df_colaborador_agregado

def gerar_atividades_por_colaborador(df,nome_colaborador):
  df_colaborador = df[df['Colaborador'] == nome_colaborador]

  df_colaborador_agregado = df_colaborador.pivot_table(
      index = ['Colaborador', 'Data', 'Atividade', 'Furo'],
      values = ['Quantidade'],
      aggfunc='sum'
  )

  df_colaborador_agregado['Quantidade'] = df_colaborador_agregado['Quantidade'].apply(converter_para_float).apply(lambda x: round(x,2))

  return df_colaborador_agregado

def gerar_dfs_atividades_por_colaborador(df):

  dfs = []

  for colaborador in df['Colaborador'].unique():
      dfs.append(
          (colaborador, gerar_atividades_por_colaborador(df, colaborador))
      )

  return dfs


def gerar_resumo_memoria_completo(dbname, user, password,host,port, initialDate, finalDate, contrato_id):
  resumo_memoria_por_colaborador = []

  df = carregar_dados_resumo_memoria(dbname, user, password, host, port, initialDate,finalDate,contrato_id)

  if df.empty:
     return df

  for colaborador in df['Colaborador'].unique():
      resumo_memoria_por_colaborador.append(
          (colaborador, gerar_memoria_de_calculo_por_colaborador(df, colaborador, dbname=dbname, user=user, password=password, host=host, port=port, contrato_id=contrato_id))
      )

  resumo_memoria_completo = pd.concat([x[1] for x in resumo_memoria_por_colaborador],axis=0)

  return resumo_memoria_completo

def obter_centro_custo_contrato(contrato_id,dbname, user, password, host, port):
  sql = f"""
    select centro_custo from contratos where id = {contrato_id}
  """

  data = execute_query(dbname,user,password,host,port,sql)

  return data[0]['centro_custo']

def exportar_para_excel(
      resumo_memoria_completo = None,
      atividades_por_colaborador = None,
      arquivo_saida = None,
      resumo_pagamento = None,
      data_inicial = None,
      data_final = None,
      contrato_id = None,
      dbname = None, user = None, password = None, host = None, port = None):

  if os.path.isfile(arquivo_saida):
    os.remove(arquivo_saida)

  def ajustar_resumo_memoria_completo():

    resumo_memoria_completo_copia = resumo_memoria_completo.copy()

    resumo_memoria_completo_copia = resumo_memoria_completo_copia.droplevel('ColaboradorId', axis=0)

    resumo_memoria_completo_copia['Quantidade'] = resumo_memoria_completo_copia['Quantidade'].round(2)


    resumo_memoria_completo_copia['Subtotal'] = resumo_memoria_completo_copia['Subtotal'].apply(converter_para_float).apply(lambda x: round(x,2))



    resumo_memoria_completo_copia.style.set_properties(**{'text-align': 'center'})

    resumo_memoria_completo_copia['Valor a pagar'] = resumo_memoria_completo_copia['Valor a pagar'].apply(lambda x : format_currency(x, currency='BRL', locale='pt_BR') if not np.isnan(x) else np.nan)

    resumo_memoria_completo_copia['Limite'] = resumo_memoria_completo_copia['Limite'].apply(lambda x : format_currency(x, currency='BRL', locale='pt_BR') if not np.isnan(x) else np.nan)

    resumo_memoria_completo_copia['Valor Unitário'] = resumo_memoria_completo_copia['Valor Unitário'].apply(lambda x : format_currency(x, currency='BRL', locale='pt_BR') if not np.isnan(x) else np.nan)

    resumo_memoria_completo_copia['Subtotal'] = resumo_memoria_completo_copia['Subtotal'].apply(lambda x : format_currency(x, currency='BRL', locale='pt_BR') if not np.isnan(x) else np.nan)


    return resumo_memoria_completo_copia

  def ajustar_df_atividades_por_colaborador(df):
    #df['Quantidade'] = df['Quantidade'].round(2)

    df['Quantidade'] = df['Quantidade'].apply(converter_para_float).apply(lambda x: round(x,2))

    #df['Quantidade'] = df['Quantidade'].astype(str).str.replace('.',',')

    return df

  def ajustar_resumo_pagamento():
     resumo_pagamento_copia = resumo_pagamento.copy()

     resumo_pagamento_copia['Valor a pagar'] = resumo_pagamento_copia['Valor a pagar'].apply(lambda x : format_currency(x, currency='BRL', locale='pt_BR') if not np.isnan(x) else np.nan)

     return resumo_pagamento_copia

  def inserir_titulo_planilha_resumo_memoria():
    wb = load_workbook(arquivo_saida)

    ws = wb['Resumo Memória']

    # ws.insert_rows(1)

    merged_cells_range = ws.merged_cells.ranges

    for merged_cell in merged_cells_range:
       merged_cell.shift(0, 1)

    ws.insert_rows(1)

    ws.merge_cells('A1:H1')

    ws['A1'].fill = PatternFill(start_color='005078', end_color='005078',fill_type='solid')

    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    centro_custo = obter_centro_custo_contrato(contrato_id, dbname, user, password, host, port)

    ws['A1'] = f'CC:_{centro_custo} MEMORIA DE CÁLCULO DA PRODUÇÃO PERÍODO {datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%d/%m/%Y")} A {datetime.strptime(data_final, "%Y-%m-%d").strftime("%d/%m/%Y")}'

    ws['A1'].font = Font(color="FFFFFF")

    wb.save(arquivo_saida)

  def inserir_titulo_planilha_resumo_pagamento():
    wb = load_workbook(arquivo_saida)

    ws = wb['Resumo Pagamento']

    ws.insert_rows(1)

    ws.merge_cells('A1:J1')

    ws['A1'].fill = PatternFill(start_color='005078', end_color='005078',fill_type='solid')

    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    centro_custo = obter_centro_custo_contrato(contrato_id, dbname, user, password, host, port)

    ws['A1'] = f'CC:_{centro_custo} RESUMO DE VALORES DE PRODUTIVIDADE REFERENTE AO PERÍODO {datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%d/%m/%Y")} A {datetime.strptime(data_final, "%Y-%m-%d").strftime("%d/%m/%Y")}'

    ws['A1'].font = Font(color="FFFFFF")

    wb.save(arquivo_saida)

  with pd.ExcelWriter(arquivo_saida, engine='xlsxwriter') as writer:

    if resumo_memoria_completo is not None:

      resumo_memoria_completo = ajustar_resumo_memoria_completo()

      resumo_memoria_completo.to_excel(
        writer,
        sheet_name='Resumo Memória'
      )

      if resumo_pagamento is not None:

        resumo_pagamento = ajustar_resumo_pagamento()

        resumo_pagamento.to_excel(writer, sheet_name='Resumo Pagamento', index=False)

    def ajustar_worksheet_resumo_memoria( format_numbers, format_center, format_border):
      worksheet_resumo_memoria = writer.sheets['Resumo Memória']

      worksheet_resumo_memoria.set_column('D:D', None, format_numbers) #quantidade
      worksheet_resumo_memoria.set_column('D:D', None, format_center) #quantidade
      #worksheet_resumo_memoria.set_column('D:D', None, format_border)

      worksheet_resumo_memoria.set_column('E:E', None, format_numbers) #valor unitario
      worksheet_resumo_memoria.set_column('E:E', None, format_center) #valor unitario
      #worksheet_resumo_memoria.set_column('E:E', None, format_border)

      worksheet_resumo_memoria.set_column('F:F', None, format_numbers) #sub total
      worksheet_resumo_memoria.set_column('F:F', None, format_center) #sub total
      #worksheet_resumo_memoria.set_column('F:F', None, format_border)

      worksheet_resumo_memoria.set_column('G:G', None, format_numbers) #total geral
      worksheet_resumo_memoria.set_column('G:G', None, format_center) #total geral
      #worksheet_resumo_memoria.set_column('G:G', None, format_border)



      worksheet_resumo_memoria.conditional_format(1, 0, len(resumo_memoria_completo) + 1, resumo_memoria_completo.index.nlevels + len(resumo_memoria_completo.columns) - 1, {'type': 'no_errors', 'format': format_border})


      worksheet_resumo_memoria.autofit()



    def ajustar_worksheet_resumo_pagamento(format_numbers, format_border):
      worksheet_resumo_pagamento = writer.sheets['Resumo Pagamento']

      colunas = ['A','B','C','D','E','F','G','H','I','J']

      # for coluna in colunas:
      #   worksheet_resumo_pagamento.set_column(f'{coluna}:{coluna}', None, format_border)

      worksheet_resumo_pagamento.autofit()

      worksheet_resumo_pagamento.set_column('K:K', None, format_numbers) #valor a pagar

      worksheet_resumo_pagamento.conditional_format(1, 0, len(resumo_pagamento) + 1, resumo_pagamento.index.nlevels + len(resumo_pagamento.columns) - 2, {'type': 'no_errors', 'format': format_border})

    workbook = writer.book


    format_numbers = workbook.add_format({'num_format': '#,##0.00'})
    # format_numbers = workbook.add_format({'num_format': '0.00'})

    format_center = workbook.add_format({'align': 'center'})

    format_border = workbook.add_format({
        'border': 1  # 1 is the thinnest border
    })

    if resumo_memoria_completo is not None:

      ajustar_worksheet_resumo_memoria(format_numbers, format_center, format_border)

    if resumo_pagamento is not None:
      ajustar_worksheet_resumo_pagamento(format_numbers, format_border)

    indice_planilha = 1

    if atividades_por_colaborador is not None:

      for nome_colaborador, df_atividades in atividades_por_colaborador:
        nome_planilha = nome_colaborador.split()[0] + f"_{indice_planilha}"

        df_atividades_ajustado = ajustar_df_atividades_por_colaborador(df_atividades)

        df_atividades_ajustado.to_excel(writer, sheet_name=nome_planilha)

        writer.sheets[nome_planilha].autofit()

        # writer.sheets[nome_planilha].set_column('E:E', None, format_border)

        writer.sheets[nome_planilha].conditional_format(
           1,
           0,
           len(df_atividades_ajustado) + 1,
            df_atividades_ajustado.index.nlevels + len(df_atividades_ajustado.columns) - 1,
            {'type': 'no_errors', 'format': format_border}
        )

        indice_planilha += 1

  if resumo_memoria_completo is not None:

    inserir_titulo_planilha_resumo_memoria()

  if resumo_pagamento is not None:
    inserir_titulo_planilha_resumo_pagamento()

def obter_resumo_pagamento_por_colaborador(colaborador_id, valorAPagar, user,password,host,port,dbname, contrato_id):





  sql = f"""
    select u.numero_cadastro as "Cadastro", u.nome as "Colaborador", c.titulo as "Cargo", u.cpf as "CPF", u.banco as "Banco", u.codigo_banco as "Código Banco", u.agencia as "Agência", u.conta as "Conta", u.pix as "PIX"
    from users u
    inner join cargos c on c.id = u.cargo_id
    where u.id = {colaborador_id}
  """

  data = execute_query(dbname,user,password,host,port,sql)

  df = pd.DataFrame(data)

  limite_pagamento = obter_limite_pagamento(colaborador_id, contrato_id, dbname, user, password, host, port)

  df['Valor a pagar'] = limite_pagamento if (limite_pagamento is not None and valorAPagar > limite_pagamento) else valorAPagar



  return df

def obter_resumo_pagamento(resumo_memoria_completo, user,password,host,port,dbname,contrato_id):

  resumos_pagamento = []

  resumo_memoria_completo_indice_resetado = resumo_memoria_completo.reset_index()

  for colaboradorId in resumo_memoria_completo_indice_resetado['ColaboradorId'].unique():
    resumos_pagamento.append(
      obter_resumo_pagamento_por_colaborador(
          colaborador_id=colaboradorId,
          valorAPagar=resumo_memoria_completo_indice_resetado[resumo_memoria_completo_indice_resetado['ColaboradorId'] == colaboradorId]['Valor a pagar'].iloc[0],
          user=user,
          password=password,
          host=host,
          port=port,
          dbname=dbname,
          contrato_id=contrato_id
      )
    )

  resumo_pagamento = pd.concat(resumos_pagamento, axis=0)

  resumo_pagamento.reset_index(drop=True, inplace=True)


  return resumo_pagamento

def executando_no_jupyter():
  return 'get_ipython' in globals()

def executando_como_script():
   return not executando_no_jupyter()

def executar_como_script():
    import argparse

    parser = argparse.ArgumentParser()

    parser.add_argument('--initialDate', required=True)
    parser.add_argument('--finalDate', required=True)
    parser.add_argument('--contractId', required=True)
    parser.add_argument('--dbname', required=True)
    parser.add_argument('--user', required=True)
    parser.add_argument('--password', required=True)
    parser.add_argument('--host', required=True)
    parser.add_argument('--port', required=True)
    parser.add_argument('--target-excel-file', required=True)
    parser.add_argument('--type-of-report', required=True, help="Valores possiveis: resumo_memoria, atividade_colaborador")
    parser.add_argument('--colaborador-id',required=False, type=int)

    args = parser.parse_args()

    if os.path.exists(args.target_excel_file):
       os.unlink(args.target_excel_file)

    if args.type_of_report == 'resumo_memoria':

      resumo_memoria_completo = gerar_resumo_memoria_completo(
          dbname=args.dbname,
          user=args.user,
          password=args.password,
          host=args.host,
          port=args.port,
          initialDate=args.initialDate,
          finalDate=args.finalDate,
          contrato_id=args.contractId
      )

      resumo_pagamento = obter_resumo_pagamento(resumo_memoria_completo,
        user=args.user,
        password=args.password,
        host=args.host,
        port=args.port,
        dbname=args.dbname,
        contrato_id=args.contractId
      )

      atividades_colaboradores = carregar_atividades_colaboradores(
          dbname=args.dbname,
          user=args.user,
          password=args.password,
          host=args.host,
          port=args.port,
          initialDate=args.initialDate,
          finalDate=args.finalDate,
          contractId=args.contractId,
          colaboradorId=args.colaborador_id
      )

      dfs_atividades_por_colaborador = None

      if len(atividades_colaboradores) > 0:

        dfs_atividades_por_colaborador = gerar_dfs_atividades_por_colaborador(atividades_colaboradores)



      exportar_para_excel(
        resumo_memoria_completo = resumo_memoria_completo,
        atividades_por_colaborador = dfs_atividades_por_colaborador,
        arquivo_saida = args.target_excel_file,
        resumo_pagamento = resumo_pagamento,
        data_inicial=args.initialDate,
        data_final=args.finalDate,
        contrato_id=args.contractId,
        dbname=args.dbname, user=args.user, password=args.password, host=args.host, port=args.port)

    if args.type_of_report == 'atividade_colaborador':

      atividades_colaboradores = carregar_atividades_colaboradores(
          dbname=args.dbname,
          user=args.user,
          password=args.password,
          host=args.host,
          port=args.port,
          initialDate=args.initialDate,
          finalDate=args.finalDate,
          contractId=args.contractId,
          colaboradorId=args.colaborador_id
      )

      if len(atividades_colaboradores) > 0:

        dfs_atividades_por_colaborador = gerar_dfs_atividades_por_colaborador(atividades_colaboradores)

        exportar_para_excel(
          resumo_memoria_completo = None,
          atividades_por_colaborador = dfs_atividades_por_colaborador,
          arquivo_saida = args.target_excel_file,
          resumo_pagamento = None,
          data_inicial=args.initialDate,
          data_final=args.finalDate,
          contrato_id=args.contractId,
          dbname=args.dbname, user=args.user, password=args.password, host=args.host, port=args.port)





#%%

if executando_no_jupyter():
   load_dotenv(r'.\dados.env',override=True)

   dados_resumo_memoria = carregar_dados_resumo_memoria(
        dbname=os.getenv('DB_NAME'),
        user=os.getenv('DB_USER'),
        password=os.getenv('DB_PASSWORD'),
        host=os.getenv('DB_HOST'),
        port=os.getenv('DB_PORT'),
        initialDate='2023-03-21',
        finalDate='2024-05-20',
        contractId=1

   )

#%%
if executando_no_jupyter():
  load_dotenv(r'.\dados.env',override=True)

  dados_resumo_memoria = carregar_dados_resumo_memoria(
      dbname=os.getenv('DB_NAME'),
      user=os.getenv('DB_USER'),
      password=os.getenv('DB_PASSWORD'),
      host=os.getenv('DB_HOST'),
      port=os.getenv('DB_PORT'),
      initialDate='2023-12-21',
      finalDate='2024-01-20',
      contractId=1

  )

  df_colaborador = dados_resumo_memoria[dados_resumo_memoria['Colaborador'] == 'Allan Henrique Hermogenes']

  df_colaborador_agregado = df_colaborador.pivot_table(
    index = ['ColaboradorId', 'Colaborador', 'Cargo', 'Atividade'],
    values=['Quantidade', 'Valor Unitário'],
    aggfunc={
        'Quantidade': 'sum',
        'Valor Unitário': 'first'
    }
  )

  df_colaborador_agregado['Subtotal'] = df_colaborador_agregado['Quantidade'] * df_colaborador_agregado['Valor Unitário']

  s = pd.Series(np.nan, index=df_colaborador_agregado.index)

  s.iloc[0] = df_colaborador_agregado['Subtotal'].sum()

  df_colaborador_agregado['Valor a pagar'] = s



#%%

if executando_no_jupyter():
    load_dotenv(r'.\dados.env',override=True)




    resumo_memoria_completo = gerar_resumo_memoria_completo(
        dbname=os.getenv('DB_NAME'),
        user=os.getenv('DB_USER'),
        password=os.getenv('DB_PASSWORD'),
        host=os.getenv('DB_HOST'),
        port=os.getenv('DB_PORT'),
        initialDate='2023-12-21',
        finalDate='2024-01-20',
        contrato_id=1

    )

    resumo_memoria_completo

#endregion

#%%

if executando_no_jupyter():
  load_dotenv(r'.\dados.env',override=True)

  initialDate = '2023-03-21'
  finalDate = '2024-04-20'
  contractId = 1

  resumo_memoria_completo = gerar_resumo_memoria_completo(
    dbname=os.getenv('DB_NAME'),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT'),
    initialDate=initialDate,
    finalDate=finalDate,
    contrato_id=contractId
  )

  atividades_colaboradores = carregar_atividades_colaboradores(
    dbname=os.getenv('DB_NAME'),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT'),
    initialDate=initialDate,
    finalDate=finalDate,
    contractId=contractId
  )

  dfs_atividades_por_colaborador = gerar_dfs_atividades_por_colaborador(atividades_colaboradores)

  resumo_pagamento = obter_resumo_pagamento(resumo_memoria_completo,
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT'),
    dbname=os.getenv('DB_NAME'),
    contrato_id=contractId
    )


  exportar_para_excel(resumo_memoria_completo, dfs_atividades_por_colaborador, 'resumo_memoria_completo.xlsx', resumo_pagamento,
                      data_inicial=initialDate, data_final=finalDate, contrato_id=contractId,
                      dbname=os.getenv('DB_NAME'), user=os.getenv('DB_USER'), password=os.getenv('DB_PASSWORD'), host=os.getenv('DB_HOST'), port=os.getenv('DB_PORT'))

  #%%

if executando_no_jupyter():

   load_dotenv(r'.\dados.env',override=True)

   resumo_pagamento_por_colaborador = obter_resumo_pagamento_por_colaborador(
    colaborador_id=223,
    valorAPagar=1000,
    dbname=os.getenv('DB_NAME'),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT'),)

#%%
if executando_no_jupyter():
    load_dotenv(r'.\dados.env',override=True)

    initialDate = '2024-03-21'
    finalDate = '2024-04-20'
    contractId = 1

    resumo_memoria_completo = gerar_resumo_memoria_completo(
      dbname=os.getenv('DB_NAME'),
      user=os.getenv('DB_USER'),
      password=os.getenv('DB_PASSWORD'),
      host=os.getenv('DB_HOST'),
      port=os.getenv('DB_PORT'),
      initialDate=initialDate,
      finalDate=finalDate,
      contrato_id=contractId
    )

    resumo_pagamento = obter_resumo_pagamento(
      resumo_memoria_completo=resumo_memoria_completo,
      user=os.getenv('DB_USER'),
      password=os.getenv('DB_PASSWORD'),
      host=os.getenv('DB_HOST'),
      port=os.getenv('DB_PORT'),
      dbname=os.getenv('DB_NAME'),
      contrato_id=contractId
    )

#%%
if executando_no_jupyter():
   query = "select limite from limite_pagamentos where contrato_id = 5 and cargo_id = 57"
   data = execute_query(
      dbname=os.getenv('DB_NAME'),
      user=os.getenv('DB_USER'),
      password=os.getenv('DB_PASSWORD'),
      host=os.getenv('DB_HOST'),
      port=os.getenv('DB_PORT'),
      query=query
   )

#%%

if executando_no_jupyter():
   from openpyxl import load_workbook
   from openpyxl.styles import PatternFill, Alignment, Font
   from pathlib import Path
   from datetime import datetime

   resumo_memoria_completo_xlsx = r'C:\projel\projel-produtividade\analysis\resumo_memoria_completo.xlsx'

   resumo_memoria_completo_xlsx_modificado = Path(resumo_memoria_completo_xlsx).parent / (Path(resumo_memoria_completo_xlsx).name[0:Path(resumo_memoria_completo_xlsx).name.index('.')] + '-' + datetime.now().strftime('%Y-%m-%d-%H-%M-%S') +  '.xlsx')

   wb = load_workbook(resumo_memoria_completo_xlsx)

   ws = wb['Resumo Memória']

   ws.insert_rows(1)

   ws.merge_cells('A1:H1')

   ws['A1'].fill = PatternFill(start_color='005078', end_color='005078',fill_type='solid')

   ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

   ws['A1'] = 'CC:_264 MEMORIA DE CÁLCULO DA PRODUÇÃO PERÍODO 21/12/2023 A 20/01/2024'

   ws['A1'].font = Font(color="FFFFFF")

   wb.save(resumo_memoria_completo_xlsx_modificado)

#%%
if executando_no_jupyter():

  import weasyprint
  import tempfile

  load_dotenv(r'.\dados.env',override=True)

  initialDate = '2023-12-21'
  finalDate = '2024-01-20'
  contractId = 1

  resumo_memoria_completo = gerar_resumo_memoria_completo(
    dbname=os.getenv('DB_NAME'),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT'),
    initialDate=initialDate,
    finalDate=finalDate,
    contrato_id=contractId
  )

  arquivo_html_temporario = tempfile.NamedTemporaryFile(delete=False, suffix='.html')

  resumo_memoria_completo.to_html(arquivo_html_temporario.name)

  weasyprint.HTML(arquivo_html_temporario.name).write_pdf('resumo_memoria_completo.pdf')

#%%
if executando_no_jupyter():
  load_dotenv(r'.\dados.env',override=True)

  initialDate = '2024-03-21'
  finalDate = '2024-05-20'
  contractId = 1

  resumo_memoria_completo = gerar_resumo_memoria_completo(
    dbname=os.getenv('DB_NAME'),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT'),
    initialDate=initialDate,
    finalDate=finalDate,
    contrato_id=contractId
  )

  resumo_pagamento = obter_resumo_pagamento(
    resumo_memoria_completo=resumo_memoria_completo,
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT'),
    dbname=os.getenv('DB_NAME'),
    contrato_id=contractId
  )
#%%
if executando_no_jupyter():
  df = resumo_memoria_completo.reset_index()
  df = df[~df['Atividade'].isin(
     ['Mudança de Sonda (Maior que 200 m)',
     'Mudança de Sonda (Menor que 200 m)'
     'Mudança de sonda (Km adicional)']
  )]

  df = df.set_index(['ColaboradorId','Colaborador','Cargo','Atividade'])



if executando_como_script():
   executar_como_script()


# %%
